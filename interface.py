import getpass
import openpyxl

def read_pins_from_txt():
    try:
        with open("pin.txt", "r") as file:
            pins = [int(line.strip()) for line in file.readlines()]
        return pins
    except FileNotFoundError:
        print("PIN file not found.")
        exit()

def write_pins_to_file(data):
    try:
        wb = openpyxl.load_workbook("database.xlsx")
        sheet = wb.active
        sheet.delete_cols(1, sheet.max_column)
        sheet.append(["PIN", "Balance", "Transaction History"])
        for entry in data:
            sheet.append([entry["PIN"], entry["Balance"], entry["Transaction History"]])
        wb.save("database.xlsx")
        wb.close()
    except FileNotFoundError:
        print("Database file not found.")
        exit()

def append_pins_to_file(pin, balance, transaction_history):
    try:
        with open("pin.txt", "a") as file:
            file.write(str(pin) + "\n")
        wb = openpyxl.load_workbook("database.xlsx")
        sheet = wb.active
        sheet.append([pin, balance, transaction_history])
        wb.save("database.xlsx")
        wb.close()
    except FileNotFoundError:
        print("Database file not found.")
        exit()

while True:
    attempts = 3
    pin_list = read_pins_from_txt()

    while attempts > 0:
        id_input = getpass.getpass(f"Please enter your 4-digit PIN (Attempts left: {attempts}): ")
        if id_input.isdigit() and 999 < int(id_input) < 10000 and int(id_input) in pin_list:
            print("PIN entered successfully!\n")
            print("Welcome! Thank you for choosing our ATM services.\n")
            print("Please choose from the below listed options: \n")
            print("0. Check Balance   1. Transaction History \n2. Withdraw        3. Deposit  \n4. Transfer        5. Reset PIN \n6. Mini Statement   7. Quit")
            break
        else:
            print("Your PIN is invalid")
            attempts -= 1
            if attempts == 0:
                print("ACCOUNT HAS BEEN LOCKED")
                exit()

    account_numbers = [123456789012, 234567890123, 345678901234, 456789012345, 67890123456, 678901234567,
                       789012345678, 890123456789, 901234567890]

    opt = int(input("\nEnter your choice~'0'/'1'/'2'/'3'/'4'/'5'/'6'/'7': "))
    bal = int(input("Enter Balance: "))
    pin_index = next((i for i, pin in enumerate(read_pins_from_txt()) if pin == int(id_input)), None)

    def switch_case(value, balance, pin_index):
        if value == 0:
            print("\nYour Current Balance is: \n", balance)
        elif value == 1:
            file = open("transactions.txt", "r")
            print(file.read())
            file.close()
        elif value == 2:
            try:
                if balance < 0:
                    print("Invalid Balance Entered")
                    return
            except ValueError:
                print("Invalid Balance Entered")
                return
            print("Withdraw: \n")
            for i in range(0, 9):
                if int(id_input) == read_pins_from_txt()[i]:
                    print("Your current balance is: ", balance)
                    withdraw = int(input("Enter the amount you want to withdraw: "))
                    if withdraw > balance:
                        print("Insufficient balance!")
                    elif 0 < withdraw <= balance:
                        print("Withdrawal successful!")
                        balance = balance - withdraw
                        print("Your new balance is: ", balance)
                        file = open("transactions.txt", "a")
                        file.write("Withdrawal: ")
                        file.write(str(withdraw))
                        file.write("\n")
                        file.close()
        elif value == 3:
            try:
                if balance < 0:
                    print("Invalid Balance")
                    return
            except ValueError:
                print("Invalid Balance")
                return
            print("Deposit: \n")
            for i in range(0, 9):
                if int(id_input) == read_pins_from_txt()[i]:
                    print("Your current balance is: ", balance)
                    deposit = int(input("Enter the amount you want to deposit: "))
                    balance = balance + deposit
                    print("Deposit successful!")
                    print("Your new balance is: ", balance)
                    file = open("transactions.txt", "a")
                    file.write("Deposit: ")
                    file.write(str(deposit))
                    file.write("\n")
                    file.close()
        elif value == 4:
            try:
                if balance < 0:
                    print("Invalid Balance")
                    return
            except ValueError:
                print("Invalid Balance")
                return
            print("Transfer: \n")
            for i in range(0, 9):
                if id_input == read_pins_from_txt()[i]:
                    print("Your current balance is: ", balance)
                    transfer = int(input("Enter the amount you want to transfer: "))
                    account = int(input("Enter the account number you want to transfer to: "))
                    if transfer > balance:
                        print("Insufficient balance!")
                    else:
                        if account in account_numbers:
                            for i in range(0, 9):
                                if account == account_numbers[i]:
                                    print("Transfer successful!")
                                    balance = balance + transfer
                                    balance = balance - transfer
                                    print(f"The amount of {transfer} has been transferred to account number: {account}")
                                    print("Transfer successful!")
                                    balance = balance - transfer
                                    print("Your new balance is: ", balance)
                                    file = open("transactions.txt", "a")
                                    file.write("Transfer: ")
                                    file.write(str(transfer))
                                    file.write("\n")
                                    file.close()
                                else:
                                    print("Invalid account number!")
        elif value == 5:
            new_pin = getpass.getpass("Enter your new 4-digit PIN: ")
            if new_pin.isdigit() and 999 < int(new_pin) < 10000:
                read_pins_from_txt()[pin_index] = int(new_pin)
                write_pins_to_file(read_pins_from_txt())
                print("PIN reset successful!")
            else:
                print("Invalid PIN format. PIN should be a 4-digit number.")
        elif value == 6:
            print_mini_statement()
        elif value == 7:
            print("Thank you for using our services!")
            exit()
        else:
            print("Invalid option! Please try again.")

    switch_case(opt, bal, pin_index)
